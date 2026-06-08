import sys
from typing import Generator
import asyncio
from pathlib import Path
import openpyxl
import datetime
from deep_translator import GoogleTranslator


PROJECT_ROOT = str(Path(__file__).resolve().parents[1])
if PROJECT_ROOT not in sys.path:
	sys.path.insert(0, PROJECT_ROOT)


async def translator(name: str) -> str:
	if name is None:
		return None
	translated = await asyncio.to_thread(
		GoogleTranslator(source="ru", target="en").translate, name.strip()
	)
	return translated.lower().replace(" ", "-").replace("_", "-")


def open_xlsx_file(file_path: str) -> Generator:
	wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	ws = wb.active
	try:
		for row in ws.iter_rows(values_only=True):
			yield list(row)
	finally:
		wb.close()


async def main(file_path: str) -> None:
	start = datetime.datetime.now()
	trans = openpyxl.Workbook()
	trans_s = trans.active
	p = 0
	for row in open_xlsx_file(file_path):
		p += 1
		try:
			row[0] = await translator(
				row[row.index(None) - 1 if row.count(None) != 0 else -1]
			)
			trans_s.append(row)
			print(p, row)
		except:  # noqa
			break
	out_path = Path(file_path).parent / "translated.xlsx"
	trans.save(str(out_path))
	print(datetime.datetime.now() - start)


if __name__ == "__main__":
	asyncio.run(main("/app/./scripts/taxonomy-with-ids.ru-RU.xlsx"))
