import sys
from typing import AsyncGenerator, Generator
import asyncio
from pathlib import Path
import openpyxl

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import delete, select

PROJECT_ROOT = str(Path(__file__).resolve().parents[1])
if PROJECT_ROOT not in sys.path:
	sys.path.insert(0, PROJECT_ROOT)


from core.db import get_db  # noqa: E402
from database.models.catalog.base import Category  # noqa: E402
from services.category_service import invalidate_categories_tree_cache  # noqa: E402


async def clear_db(db_session: AsyncSession) -> True:
	await db_session.execute(delete(Category))
	await db_session.commit()
	return True


async def add_root_category(db_session: AsyncGenerator) -> None:
	root_category: Category = Category(
		name="Все товары",
		slug="all",
		description="Все товары",
		parent_id=None,
		is_active=True,
	)
	db_session.add(root_category)
	await db_session.commit()
	await db_session.refresh(root_category)


async def slug_generator(slug: str, db_session: AsyncGenerator) -> str:
	result = await db_session.execute(select(Category).where(Category.slug == slug))
	result = result.scalar_one_or_none()

	if result:
		result = await db_session.execute(
			select(Category).where(Category.slug.like(f"{slug}(%)"))
		)
		result_obj = result.scalars().all()

		if result_obj == []:
			slug += "(1)"
		else:
			slug += "(" + str(int((result_obj[-1].slug)[len(slug) + 1 : -1]) + 1) + ")"
	return slug


async def add_category_in_db(path: list, slug: str, db_session: AsyncSession) -> bool:
	parent_id = None

	for parent_name in path[0 : len(path) - 1]:
		result = await db_session.execute(
			select(Category).where(
				Category.name == parent_name, Category.parent_id == parent_id
			)
		)
		parent_obj = result.scalar_one_or_none()
		if parent_obj is not None:
			parent_id = parent_obj.id
		else:
			print(f"error: {path}")
			return False

	name = path[-1]
	slug = await slug_generator(slug, db_session)
	category: Category = Category(
		name=name, slug=slug, description=name, parent_id=parent_id, is_active=True
	)
	db_session.add(category)
	await db_session.commit()
	return True


def open_xlsx_file(file_path: str) -> Generator:
	wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	ws = wb.active
	try:
		for row in ws.iter_rows(values_only=True):
			yield list(row)
	finally:
		wb.close()


async def category_parser(db_session: AsyncSession, file_path: str) -> bool:
	await add_root_category(db_session)
	p = 0
	wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	max_row = (wb.active).max_row
	wb.close()
	for row in open_xlsx_file(file_path):
		p += 1
		print(f"{p}/{max_row} {row}")
		slug = row[0]
		row[0] = "Все товары"
		row = row[0 : row.index(None) if row.count(None) != 0 else len(row)]
		await add_category_in_db(row, slug, db_session)

	return True


async def main() -> None:
	db_gen: AsyncGenerator[AsyncSession, None] = get_db()
	db_session: AsyncSession = await db_gen.__anext__()
	await clear_db(db_session)

	await category_parser(db_session, "/app/./scripts/translated.xlsx")
	await invalidate_categories_tree_cache(db_session)
	return True


if __name__ == "__main__":
	asyncio.run(main())
